"""
甍AI積算エンジン — 見積書Excel出力 v2
Evidence ID連携版

摘要欄: RULE-xxx / EV-P2026-001-xxx
これで 図面 → 赤ペン → Evidence → 見積書 が一本線になる
"""
import json
from copy import copy
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Evidence生成（AssemblyからEvidence IDと根拠を作る）──
def build_evidence_map(assemblies: dict, project_id: str) -> dict:
    """
    Assembly結果からEvidence辞書を生成
    戻り値: {evidence_id: Evidence}
    """
    ev_map = {}
    seq = 1

    def mk_ev(item_name, value, unit, formula, rule_refs, calc_base,
              slope_rate=None, case_ref=None, confidence=90, human_req=False, note=""):
        nonlocal seq
        ev_id = f"EV-{project_id}-{seq:03d}"
        seq += 1
        ev = {
            "id": ev_id,
            "projectId": project_id,
            "itemName": item_name,
            "value": value,
            "unit": unit,
            "formula": formula,
            "ruleRefs": rule_refs if isinstance(rule_refs, list) else [rule_refs],
            "calcBase": calc_base,
            "slopeRate": slope_rate,
            "caseRef": case_ref,
            "confidence": confidence,
            "humanRequired": human_req,
            "correctionNote": note or None,
            "version": "Ver.3.0"
        }
        ev_map[ev_id] = ev
        return ev_id

    a = assemblies

    # 屋根工事 Evidence
    if "Eave" in a:
        e = a["Eave"]
        a["Eave"]["_ev_noki"]    = mk_ev("軒先（桟鼻・捨唐草60）", e["length"], "m",
            f"立面図実長", ["RULE-002","RULE-003"], "elevation_actual", confidence=90)

    if "SnowStop" in a:
        ss = a["SnowStop"]
        a["SnowStop"]["_ev"] = mk_ev("雪止め金具", ss["length"], "m",
            "軒先長さと同一（水下のみ）", ["RULE-002"], "elevation_actual", confidence=88)

    if "Ridge" in a:
        rd = a["Ridge"]
        a["Ridge"]["_ev_ridge"] = mk_ev("片棟（通気部材）", rd["length"], "m",
            "立面図実長 勾配伸び率不要", ["RULE-002","RULE-003","RULE-005"], "elevation_actual", confidence=88)
        if rd.get("ventCount", 0) > 0:
            a["Ridge"]["_ev_vent"] = mk_ev("換気部材（片流れ換気）", rd["ventCount"], "本",
                f"天井面積÷17.5=切上", ["RULE-005","RULE-201"], "ceiling_area", confidence=85)

    if "Apron" in a:
        ap = a["Apron"]
        a["Apron"]["_ev"] = mk_ev("雨押え板金", ap["length"], "m",
            "立面図実長 伸び率不要", ["RULE-003","RULE-007"], "elevation_actual",
            confidence=ap.get("confidence",70),
            human_req=ap.get("humanRequired",False))

    if "Vent" in a:
        vc = a["Vent"]
        if vc.get("amaoshiVentCount", 0) > 0:
            a["Vent"]["_ev_ama"] = mk_ev("雨押え換気（下屋）", vc["amaoshiVentCount"], "本",
                "下屋天井面積÷18.4=切上（2F含まない）", ["RULE-004","RULE-202"], "shimoya_ceiling_only",
                confidence=80)

    if "Valley" in a:
        vl = a["Valley"]
        a["Valley"]["_ev"] = mk_ev("谷板金（発注）", vl["orderLength"], "m",
            f"実長{vl['actualLength']}m÷有効1.9m=ceil({vl['pieces']}枚)×2m",
            ["RULE-120","RULE-121","RULE-122"], "valley_line",
            confidence=vl.get("confidence", 90))

    # 雨樋工事 Evidence
    if "Gutter" in a:
        gt = a["Gutter"]
        a["Gutter"]["_ev"] = mk_ev("軒とい NF-I型", gt["length"], "m",
            "立面図実長（水下軒先）", ["RULE-002","RULE-300"], "elevation_actual", confidence=90)

    if "Downspout" in a:
        ds = a["Downspout"]
        a["Downspout"]["_ev_kasui"] = mk_ev("集水器 F型", ds["count"], "か所",
            f"建物角{ds['count']}か所 投影面積÷69㎡確認", ["RULE-006","RULE-008"], "projection_area",
            confidence=88)
        a["Downspout"]["_ev_tate"] = mk_ev("たてとい φ60", ds["totalLength"], "m",
            f"軒高{ds['height']}m×{ds['count']}本 実長のみ",
            ["RULE-011"], "elevation_actual", confidence=88)
        a["Downspout"]["_ev_pmasu"] = mk_ev("Pマス", ds["pmasuCount"], "か所",
            "たてとい本数と同数", ["RULE-008"], "drainage_count", confidence=90)

    return ev_map


def build_rows_with_evidence(assemblies: dict) -> dict:
    """行データにEvidence IDを含める"""
    PRICE = {
        "桟鼻":750, "捨唐草60":1000, "捨唐草45":900, "雪止め":2400,
        "通気部材":400, "換気2P":27000, "換気1P":23000,
        "雨押え":2000, "水上":1700, "破風":2800, "谷板金":5000,
        "諸経費屋根":75000,
        "軒とい":3850, "集水器":2950, "たてとい":2200, "Pマス":2000,
        "諸経費雨樋":15000,
    }
    a = assemblies
    roof, gutter = [], []

    # (品目名, 数量, 単位, 単価, RuleID, EvidenceID)
    if "Eave" in a:
        e = a["Eave"]
        ev = e.get("_ev_noki","")
        roof.append(("桟鼻（後付け）",        e["length"],"m",   PRICE["桟鼻"],   "RULE-002",ev))
        roof.append(("捨て唐草60（軒）",       e["length"],"m",   PRICE["捨唐草60"],"RULE-002",ev))

    if "Verge" in a and a["Verge"]["length"] > 0:
        vg = a["Verge"]
        ev = vg.get("_ev","")
        roof.append(("捨て唐草45（妻・片棟）", vg["length"],"m",  PRICE["捨唐草45"],"RULE-003",ev))

    if "SnowStop" in a:
        ss = a["SnowStop"]
        ev = ss.get("_ev","")
        roof.append(("雪止め金具（シングル）", ss["length"],"m",   PRICE["雪止め"], "RULE-002",ev))

    if "Ridge" in a:
        rd = a["Ridge"]
        ev_rd = rd.get("_ev_ridge",""); ev_vt = rd.get("_ev_vent","")
        roof.append(("通気部材取付施工費",      rd["length"],"m",  PRICE["通気部材"],"RULE-005",ev_rd))
        if rd.get("ventCount",0) > 0:
            roof.append(("換気部材 I-ROOF 2P", rd["ventCount"],"ヶ所",PRICE["換気2P"],"RULE-201",ev_vt))

    if "Vent" in a and a["Vent"].get("amaoshiVentCount",0) > 0:
        vc = a["Vent"]
        ev = vc.get("_ev_ama","")
        roof.append(("雨押え換気 I-ROOF 2P",  vc["amaoshiVentCount"],"ヶ所",PRICE["換気2P"],"RULE-202",ev))

    if "Apron" in a and a["Apron"]["length"] > 0:
        ap = a["Apron"]
        ev = ap.get("_ev","")
        flag = "  ⚠要確認" if ap.get("humanRequired") else ""
        roof.append(("雨押え板金（平行壁）",   ap["length"],"m",  PRICE["雨押え"],  f"RULE-003{flag}",ev))

    # 破風は常に入力待ち（RULE-010: 鼻隠し+ケラバ+片棟統合なので手確認）
    roof.append(("破風板金（特注加工）",        None,"m",          PRICE["破風"],   "RULE-010",""))

    if "Valley" in a:
        vl = a["Valley"]
        ev = vl.get("_ev","")
        roof.append(("谷板金 本体",            vl["orderLength"],"m",PRICE["谷板金"],"RULE-121",ev))

    roof.append(("諸経費",                     1,"式",            PRICE["諸経費屋根"],"",""))

    # 雨樋
    if "Gutter" in a:
        gt = a["Gutter"]
        ev = gt.get("_ev","")
        gutter.append(("軒とい　ファインスケアNF-1型", gt["length"],"m",PRICE["軒とい"],"RULE-002",ev))

    if "Downspout" in a:
        ds = a["Downspout"]
        gutter.append(("集水器　F型",          ds["count"],"ヶ所", PRICE["集水器"],"RULE-008",ds.get("_ev_kasui","")))
        gutter.append(("たてとい",             ds["totalLength"],"m",PRICE["たてとい"],"RULE-011",ds.get("_ev_tate","")))
        gutter.append(("P型集水器（Pマス）",    ds["pmasuCount"],"ヶ所",PRICE["Pマス"],"RULE-008",ds.get("_ev_pmasu","")))

    gutter.append(("諸経費",                   1,"式",            PRICE["諸経費雨樋"],"",""))

    return {"屋根工事":roof, "雨樋工事":gutter}


def write_mitsumori_v2(template_path, out_path, project_info, rows_data, ev_map, project_id):
    """甍フォーマットにEvidence ID付きで書き込む"""
    wb = load_workbook(template_path)

    def safe_set(ws, row, col, value, bold=False, color=None, size=None):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell): return
        cell.value = value
        if bold or color or size:
            f = cell.font
            cell.font = Font(
                name=f.name or "メイリオ",
                size=size or f.size,
                bold=bold if bold else f.bold,
                color=color or (f.color.rgb if f.color and hasattr(f.color,'rgb') else "000000")
            )

    # 確信度に応じたセル背景色
    CONF_FILL = {
        "high":  PatternFill("solid", start_color="E2EFDA"),  # 緑 ≥80%
        "mid":   PatternFill("solid", start_color="FFF2CC"),  # 黄 70-79%
        "low":   PatternFill("solid", start_color="FFCCCC"),  # 赤 <70%
    }

    def conf_fill(confidence):
        if confidence >= 80: return CONF_FILL["high"]
        if confidence >= 70: return CONF_FILL["mid"]
        return CONF_FILL["low"]

    def write_sheet(ws, sheet_type, rows):
        safe_set(ws, 5,  16, project_info["client"])
        safe_set(ws, 9,  3,  project_info["name"])
        safe_set(ws, 10, 3,  project_info["address"])
        safe_set(ws, 11, 3,  sheet_type)

        for i, row_data in enumerate(rows[:13]):
            r = 18 + i
            item, qty, unit, tanka, rule_id, ev_id = row_data

            safe_set(ws, r, 1,  item)  # A18:O18マージセル先頭はA列(col=1)
            if qty is not None:
                safe_set(ws, r, 16, qty)
            safe_set(ws, r, 19, unit)
            if tanka is not None:
                safe_set(ws, r, 21, tanka)

            # 摘要: RULE-xxx / EV-xxx （RULE-402準拠）
            tekiyo = rule_id
            if ev_id:
                tekiyo += f" / {ev_id}"
            safe_set(ws, r, 29, tekiyo)

            # Confidence表示: Evidence IDがあればセルに色付け
            if ev_id and ev_id in ev_map:
                ev = ev_map[ev_id]
                conf = ev.get("confidence", 90)
                fill = conf_fill(conf)
                for c in [2, 16, 19, 21]:
                    cell = ws.cell(r, c)
                    if not isinstance(cell, MergedCell):
                        cell.fill = fill
                # 確信度を品目欄末尾に追記
                cell_b = ws.cell(r, 2)
                if not isinstance(cell_b, MergedCell):
                    if ev.get("humanRequired"):
                        cell_b.font = Font(name="メイリオ", color="C00000", bold=True)

    # 屋根工事シート
    ws1 = wb["Sheet1"]
    write_sheet(ws1, "屋根工事", rows_data["屋根工事"])
    ws1.title = "御見積書（屋根工事）"

    # 雨樋工事シートをコピー
    ws2 = wb.copy_worksheet(ws1)
    ws2.title = "御見積書（雨樋工事）"
    for r in range(18, 31):
        for c in [1, 16, 19, 21, 29]:
            cell = ws2.cell(row=r, column=c)
            if not isinstance(cell, MergedCell):
                cell.value = None
    write_sheet(ws2, "雨樋工事", rows_data["雨樋工事"])

    # Evidence一覧シートを追加
    ws3 = wb.create_sheet("Evidence一覧")
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 22
    ws3.column_dimensions['C'].width = 10
    ws3.column_dimensions['D'].width = 6
    ws3.column_dimensions['E'].width = 28
    ws3.column_dimensions['F'].width = 22
    ws3.column_dimensions['G'].width = 10
    ws3.column_dimensions['H'].width = 10

    navy = "1F3864"
    ws3.cell(1,1,"Evidence一覧（積算根拠トレース）").font = Font(name="メイリオ",bold=True,size=12,color="FFFFFF")
    ws3.cell(1,1).fill = PatternFill("solid",start_color=navy)
    ws3.merge_cells("A1:H1")
    ws3.row_dimensions[1].height = 24

    hdrs = ["Evidence ID","項目名","数量","単位","算出根拠","RuleID参照","Conf%","人確認"]
    for c, h in enumerate(hdrs, 1):
        cell = ws3.cell(2, c, h)
        cell.font = Font(name="メイリオ",bold=True,color="FFFFFF",size=10)
        cell.fill = PatternFill("solid",start_color="2E75B6")
        cell.alignment = Alignment(horizontal="center",vertical="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
    ws3.row_dimensions[2].height = 18

    for r_idx, (ev_id, ev) in enumerate(ev_map.items(), 3):
        ws3.row_dimensions[r_idx].height = 18
        conf = ev.get("confidence",90)
        fill = conf_fill(conf)
        bg   = "FFFFFF" if r_idx % 2 == 0 else "F5F8FF"
        thin = Side(style="thin",color="AAAAAA")
        bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)
        row_vals = [
            ev_id,
            ev["itemName"],
            ev["value"],
            ev["unit"],
            ev.get("formula",""),
            ", ".join(ev.get("ruleRefs",[])),
            conf,
            "⚠要確認" if ev.get("humanRequired") else "✅OK",
        ]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws3.cell(r_idx, c_idx, val)
            cell.fill = fill if c_idx in [7,8] else PatternFill("solid",start_color=bg)
            cell.border = bdr
            cell.font = Font(name="メイリオ",size=10,
                color="C00000" if (c_idx==8 and ev.get("humanRequired")) else "375623" if c_idx==7 and conf>=80 else "000000")
            cell.alignment = Alignment(
                horizontal="right" if c_idx in [3,7] else "center" if c_idx in [4,8] else "left",
                vertical="center"
            )
            if c_idx == 3 and isinstance(val,(int,float)):
                cell.number_format = "#,##0.00"

    wb.save(out_path)
    return out_path


# ── 実行 ──
if __name__ == "__main__":
    PROJECT_ID = "P2026-001"

    assemblies = {
        "Eave":      {"type":"Eave",     "length":19.0, "pieces":11, "ruleRefs":["RULE-002","RULE-003"]},
        "Verge":     {"type":"Verge",    "length":0.0,  "pieces":0,  "ruleRefs":["RULE-002","RULE-003"]},
        "SnowStop":  {"type":"SnowStop", "length":19.0, "units":19,  "ruleRefs":["RULE-002"]},
        "Ridge":     {"type":"Ridge",    "length":16.6, "pieces":10, "ventCount":1,
                      "ruleRefs":["RULE-002","RULE-005","RULE-201"]},
        "Vent":      {"type":"Vent",     "mainVentCount":1,"amaoshiVentCount":0,"totalCount":1,
                      "ruleRefs":["RULE-004","RULE-201"]},
        "Apron":     {"type":"Apron",    "length":0.0,  "pieces":0,  "humanRequired":False,
                      "confidence":90,   "ruleRefs":["RULE-003"]},
        "Gutter":    {"type":"Gutter",   "length":19.0, "pieces":5,  "gutterCount":2,
                      "areaPerGutter":42.8,"ruleRefs":["RULE-002","RULE-300"],"alerts":[]},
        "Downspout": {"type":"Downspout","height":6.475,"count":2,   "totalLength":12.95,
                      "elbows":4,"pmasuCount":2,"ruleRefs":["RULE-011"]},
    }

    # Evidence生成
    ev_map = build_evidence_map(assemblies, PROJECT_ID)
    print(f"Evidence生成: {len(ev_map)}件")
    for eid, ev in ev_map.items():
        hr = "⚠" if ev["humanRequired"] else "✅"
        print(f"  {eid}: {ev['itemName']} = {ev['value']}{ev['unit']}  [{','.join(ev['ruleRefs'])}]  Conf.{ev['confidence']}%  {hr}")

    # 行データ生成
    rows = build_rows_with_evidence(assemblies)

    print("\n屋根工事 明細:")
    for r in rows["屋根工事"]:
        item,qty,unit,tanka,rule,ev = r
        q = f"{qty}" if qty is not None else "─"
        ev_short = ev[-7:] if ev else "  ─  "
        print(f"  {item:<28} {q:>7} {unit:<5} ¥{tanka:>6,}  {rule:<18}  {ev_short}")

    # Excel出力
    out = write_mitsumori_v2(
        template_path="/tmp/mitsumori.xlsx",
        out_path="/mnt/user-data/outputs/関根様邸_御見積書_Evidence連動.xlsx",
        project_info={"client":"関根　柊平","name":"関根　柊平様邸 新築工事","address":"練馬区南大泉3-426-15"},
        rows_data=rows,
        ev_map=ev_map,
        project_id=PROJECT_ID
    )

    # AnnotationJSONにEvidenceを統合して保存
    try:
        with open("/mnt/user-data/outputs/関根様邸_AnnotationJSON.json", encoding="utf-8") as f:
            ann_data = json.load(f)
        ann_data["evidences"] = ev_map
        with open("/mnt/user-data/outputs/関根様邸_AnnotationJSON.json","w",encoding="utf-8") as f:
            json.dump(ann_data, f, ensure_ascii=False, indent=2)
        print("\nAnnotationJSON にEvidence統合完了")
    except Exception as e:
        print(f"\nAnnotationJSON更新スキップ: {e}")

    print(f"\n完了: {out}")
